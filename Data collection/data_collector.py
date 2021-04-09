from __future__ import absolute_import
import dijkstra_alg
import ftsp_alg
import dijkstra_data
import ftsp_data
import openpyxl

def room_list_generator(road_network):
    room_list = []
    for row in xrange(0, len(road_network)):
        for col in xrange(0,len(road_network[row])):
            current_vertex =road_network[row][col].lower()
            if current_vertex != "none":
                # checks if vertex is already in room_list. If it isn't, add the new entry
                if not any( current_vertex == vertex for vertex in room_list):
                    room_list.append(current_vertex)
    return room_list

def map_import():
    wb = openpyxl.load_workbook("road network map.xlsx")  # Work Book
    ws = wb["Map"] # Work Sheet
    road_network = []
    for row in ws.iter_rows(min_row=1):
        temp = []
        for x in xrange(0,len(row)):
            temp.append(unicode(row[x].value).lower())
        road_network.append(temp)
    return road_network

def data_export(data):
    wb = openpyxl.load_workbook("data.xlsx")  # Work Book
    ws = wb["Sheet1"] # Work Sheet
    ws.cell(row = 1, column = 1).value  = "Start location"
    ws.cell(row = 1, column = 2).value  = "Target location"
    ws.cell(row = 1, column = 3).value  = "Dijkstra number of rooms in path"
    ws.cell(row = 1, column = 4).value  = "Dijkstra number of turns in path"
    ws.cell(row = 1, column = 5).value = "ftsp number of rooms in path"
    ws.cell(row = 1, column = 6).value  = "ftsp number of turns in path"
    current_row = 1
    for data_row in data:
        current_row +=1
        current_col = 1
        for value in data_row:
            ws.cell(row = current_row, column = current_col).value = value
            current_col +=1
    wb.save("data.xlsx")

def main():

    road_network = map_import()
    room_list = room_list_generator(road_network)
    data = []
    for start in room_list:
        if "desk" in start:
            for room in room_list:
                if room != start and "floor" not in room:
                    d_path = dijkstra_alg.main(road_network,start,room)
                    d_room_num, d_turn_num = dijkstra_data.main(road_network,d_path)

                    # print(d_room_num)
                    # print(d_turn_num)
                    p_path = ftsp_alg.main(road_network,start,room)
                    # print(p_path)
                    p_room_num, p_turn_num, p_path= ftsp_data.main(road_network,p_path)
                    # print(p_room_num)
                    # print(p_turn_num)
                    data.append([start,room,d_room_num,d_turn_num,p_room_num,p_turn_num])
                    data_export(data)

main()
