from __future__ import absolute_import
import ftsp_alg
import openpyxl
from naoqi import ALProxy
import speech_recognition
import pyaudio

def find_path(road_network,current_vertex,target_vertex):
    sub_path =[]
    indexes = []
    #find indexes of current_vertex
    for row in xrange(0, len(road_network)):
        for col in xrange(0,len(road_network[row])):
            if road_network[row][col] == current_vertex:
                indexes.append([row,col])

    directions = [[0,-1],[-1,0],[0,1],[1,0]]#left up right down
    #for each instance of current_vertex go in all four directions to see if there is a path to the target vertex
    for index in indexes:
        for direction in directions:
            row = index[0]
            col = index[1]
            current_vertex = road_network[row][col]
            sub_path= [current_vertex]
            #checks if the current vertex is not a empty cell
            while current_vertex != "none":
                row += direction[0]
                col += direction[1]
                #checks if the coordinates are valid
                if row < len(road_network) and row>=0:
                    if col <len(road_network[0]) and col>=0:
                        current_vertex = road_network[row][col]
                        #if the current vertex is not an empty cell and isn't already the same as the starting room or in the path
                        #this removes duplicate instances of the same room name
                        if current_vertex != "none" and current_vertex != road_network[index[0]][index[1]] and not any(current_vertex == vertex1 for vertex1 in sub_path):
                            sub_path.append(current_vertex)
                        if sub_path[-1] == target_vertex:
                            del sub_path[0]
                            return sub_path
                    else:
                        current_vertex = "none"
                else:
                    current_vertex = "none"

    return None

def room_list_generator(road_network):
    room_list = []
    #for each cell in the excel map
    for row in xrange(0, len(road_network)):
        for col in xrange(0,len(road_network[row])):
            current_vertex = road_network[row][col]
            if current_vertex != "none":
                # checks if vertex is already in room_list. If it isn't, add the new entry
                if not any( current_vertex == vertex for vertex in room_list):
                    room_list.append(current_vertex)
    return room_list

def path_filler(road_network,path):
    new_path = []
    #fills in the path with all the rooms it goes through
    for index in xrange(0,len(path)-1):
        current = path[index]
        next_room = path[index+1]
        sub_path = find_path(road_network,current,next_room)
        if index ==0:
            new_path.append(current)
        for room in sub_path:
            new_path.append(room)
    temp_new_path = new_path[:]
    safe_index = None
    # deletes the extra rooms for floors and stairs
    for room_index in xrange(0,len(temp_new_path)-2):
        if "floor" in temp_new_path[room_index] and room_index != safe_index:
            start_floor = temp_new_path[room_index]
            end_floor = temp_new_path[room_index+2]
            if start_floor[:start_floor.rindex(' ')] != end_floor[:end_floor.rindex(' ')] and "stairs" in temp_new_path[room_index+1]:
                new_path.remove(temp_new_path[room_index+1])
            else:
                safe_index = room_index +2
    return new_path

def find_direction(road_network,current_vertex,next_vertex):
    indexes = []
    #find indexes of current_vertex
    for row in xrange(0, len(road_network)):
        for col in xrange(0,len(road_network[row])):
            if road_network[row][col] == current_vertex:
                indexes.append([row,col])

    directions = [[0,-1],[-1,0],[0,1],[1,0]]#left up right down
    #for each instance of current_vertex go in all four directions to see if there is a path to the target vertex
    for index in indexes:
        for direction in directions:
            row = index[0]
            col = index[1]
            current_vertex = road_network[row][col]
            #checks if the current vertex is not a empty cell
            while current_vertex != "none" and current_vertex != next_vertex:
                row += direction[0]
                col += direction[1]
                #checks if the coordinates are valid
                if row < len(road_network) and row>=0:
                    if col <len(road_network[0]) and col>=0:
                        current_vertex = road_network[row][col]
                        #finds direction of path
                        if current_vertex == next_vertex:
                            if direction == [0,-1]:
                                return 0
                            elif direction == [-1,0]:
                                return 1
                            elif direction == [0,1]:
                                return 2
                            elif direction == [1,0]:
                                return 3
                    else:
                        current_vertex = "none"
                else:
                    current_vertex = "none"

def instructions_printer(start_location,path,road_network,instruct_type,tts,easy_access):
    full_path=path_filler(road_network,path)
    prev_used_stairs = False
    #for each room in the turning point path
    for index in xrange(0,len(path)-1):
        current = path[index]
        next_room = path[index+1]
        #finds path between turning points
        sub_path = full_path[full_path.index(current):full_path.index(next_room)]
        stairs = False

        #finds out if there is stairs are used between turning points
        for room in xrange(0,len(sub_path)):
            if "stairs" in sub_path[room]:
                end_floor = full_path[full_path.index(sub_path[room])+1]
                stair_string = full_path[full_path.index(sub_path[room])]
                if easy_access == "yes":
                    stair_string = stair_string[:stair_string.index('stairs')] + "lift"

                start_floor =  full_path[full_path.index(sub_path[room])-1]
                stairs = True

        prev_direction = find_direction(road_network,full_path[full_path.index(current)-1],current)
        direction = find_direction(road_network,current,next_room)
        if easy_access == "yes":
            transport_type = "lift"
            if "stairs" in current:
                current = current[:current.index('stairs')] + "lift"
            if "stairs" in next_room:
                next_room = next_room[:next_room.index('stairs')] + "lift"
        else:
            transport_type = "stairs"
        if instruct_type == "survey":
            if direction == 0:
                string_direction = "west"
            elif direction == 1:
                string_direction = "north"
            elif direction == 2:
                string_direction = "east"
            elif direction == 3:
                string_direction = "south"
            if stairs:
                if int(start_floor[start_floor.rindex(' '):])> int(end_floor[end_floor.rindex(' '):]):
                  string_stair_direction = "down"
                else:
                  string_stair_direction = "up"
            #adjusts name when about to use stairs
            if prev_used_stairs == False:
                if 'floor' in current:
                    current = current[:current.index('floor')] + transport_type
                if 'floor' in next_room:
                    next_room = next_room[:next_room.index('floor')] + transport_type
            else:
                prev_used_stairs = False
            tts.say(str("Face " + string_direction + " at " + current))
            if stairs:
                tts.say(str("Go " + string_stair_direction + " the "+ transport_type+ " to floor" +  end_floor[end_floor.rindex(' '):]))
                prev_used_stairs = True
            else:
                tts.say(str("Go straight until you reach " + next_room))

        elif instruct_type == "route":
            # cannot calculate change of direction for first room
            if index == 0:
                if stairs:
                    #adjusts name when about to use stairs
                    if prev_used_stairs == False:
                        if 'floor' in current:
                            current = current[:current.index('floor')] + transport_type
                        if 'floor' in next_room:
                            next_room = next_room[:next_room.index('floor')] + transport_type
                    else:
                        prev_used_stairs = False
                    tts.say(str("Go straight until you reach " + stair_string))
                    if int(start_floor[start_floor.rindex(' '):])> int(end_floor[end_floor.rindex(' '):]):
                        string_direction = "down"
                    else:
                        string_direction = "up"
                    tts.say(str("Go " + string_direction + " the " + transport_type +" to floor" +  end_floor[end_floor.rindex(' '):]))
                    prev_used_stairs = True
                else:
                    if 'floor' in next_room:
                        next_room = next_room[:next_room.index("floor")] + transport_type
                    tts.say(str("Go straight until you reach " + next_room))
            else:
                change_direction = 0
                temp = prev_direction
                while temp!=direction:
                    if temp==3:
                        temp=0
                    else:
                        temp+=1
                    change_direction+=1
                if change_direction == 1:
                    string_direction = "right"
                elif change_direction == 2:
                    string_direction = "around"
                elif change_direction == 3:
                    string_direction = "left"
                if stairs:
                    if int(start_floor[start_floor.rindex(' '):])> int(end_floor[end_floor.rindex(' '):]):
                        string_stair_direction = "down"
                    else:
                        string_stair_direction = "up"
                #adjusts name when about to use stairs
                if prev_used_stairs == False:
                    if 'floor' in current:
                        current = current[:current.index('floor')] + transport_type
                    if 'floor' in next_room:
                        next_room = next_room[:next_room.index('floor')] + transport_type
                else:
                    prev_used_stairs = False
                tts.say(str("Turn " + string_direction + " at " + current))
                if stairs:
                    tts.say(str("Go " + string_stair_direction + " the "+ transport_type + " to floor" +  end_floor[end_floor.rindex(' '):]))
                    prev_used_stairs = True
                else:
                    tts.say(str("Go straight until you reach " + next_room))

def map_import():
    wb = openpyxl.load_workbook("road network map.xlsx")
    ws = wb["Map"]
    road_network = []
    #add every cell to the road_network
    for row in ws.iter_rows(min_row=1):
        temp = []
        for col in xrange(0,len(row)):
            temp.append(unicode(row[col].value).lower())
        road_network.append(temp)
    return road_network

def map_adjust(road_network,easy_access):
    new_road_network = road_network[:]
    wb = openpyxl.load_workbook("road network map.xlsx")
    ws = wb["Rooms"]
    deleting_list = []
    #finds rooms that need to be deleted
    for current_row in ws.iter_rows(min_row=2):
        status = current_row[1].value
        access = current_row[2].value
        if status.lower() == "closed" or (easy_access == "yes" and access.lower() == "no"):
            deleting_list.append(current_row[0].value)

    #deletes rooms in the deleting list
    for row in xrange(0, len(road_network)):
        for col in xrange(0,len(road_network[row])):
            current_room = road_network[row][col].lower()
            for room in deleting_list:
                if current_room == room:
                    new_road_network[row][col] = 'none'
    return new_road_network

def config_import():
    configuration_file = open("configuration.txt")
    data_in_file = configuration_file.readlines()
    ip_address = data_in_file[0][:data_in_file[0].index(':')-1]
    port_number = data_in_file[0][data_in_file[0].index(':')+1:]
    port_number = port_number.strip('\n')
    start = data_in_file[1]
    start = start.strip('\n')
    configuration_file.close()
    return ip_address, port_number, start

def find_room_input(mic,recogniser,tts):
    road_network = map_import()
    room_list = room_list_generator(road_network)
    no_room_input = True
    target = None
    while no_room_input:
        no_input = True
        while no_input:
            with mic as source:
                audio_text = recogniser.listen(source)
                target = recogniser.recognize_google(audio_text)
                #checks if an input is given
                if target != None:
                    #checks if the input matches any of the questions
                    target = target.lower()
                    words = target.split()
                    questions = [["how","do","i","get","to"],["where","is","the"],["how","do","i","see"]]
                    is_question = False
                    condition = True
                    question_index = 0
                    word_index = 0

                    while condition:
                        question = questions[question_index]
                        if words[word_index] == question[word_index]:
                            is_question = True
                        else:
                            is_question = False
                        if is_question:
                            if (word_index < len(question)-1) and (word_index <len(words)-1) :
                                word_index +=1
                            else:
                                condition = False
                        else:
                            if question_index <len(questions)-1:
                                question_index+=1
                                word_index = 0
                            else:
                                condition = False
                                is_question = False
                    #if the input is a question
                    question_index = None
                    if is_question:
                        target_list = words[len(questions[questions.index(question)]):]
                        target = ""
                        for word in target_list:
                            target = target +" " +  word
                        target = target[1:]
                        if question == questions[0]:
                            #check if the room in the input is valid
                            if any(target == room for room in room_list):
                                question_index = 0
                            else:
                                tts.say("Sorry, that is an invalid room name. Please try again.")
                        elif question == questions[1] or question == questions[2]:
                            question_index = 1
                    else:
                        tts.say("Sorry, I didn't understand that request")

            #confirms the room name
            if question_index == 0:
                no_input = True
                while no_input:
                    with mic as source:
                        tts.say(str("Do you want to go to " + target + "?"))
                        audio_text = recogniser.listen(source)
                    try:
                        confirm = recogniser.recognize_google(audio_text)
                        if confirm == "yes":
                            no_input = False
                            no_room_input = False
                        elif confirm == "no":
                            no_input = False
                            tts.say("Oh ok")
                        else:
                            tts.say("Sorry, I don't understand. Please try again.")
                    except:
                        tts.say("Sorry, can you repeat that?")

            #finds the room name through asking keywords
            elif question_index == 1:
                wb = openpyxl.load_workbook("road network map.xlsx")  # Work Book
                ws = wb["Rooms"]
                targets = target.split()
                #for each row in the excel rooms sheet
                for current_row in ws.iter_rows(min_row=2):
                    art_keywords = []
                    #finds the keywords in the row
                    for col in range(3,len(current_row)):
                        if current_row[col].value != None:
                            art_keywords.append(current_row[col].value)
                    #compares the keywords found to the keyword from the input
                    for art_keyword in art_keywords:
                        if target in art_keyword.lower():
                            no_keyword_confirm_input = True
                            no_input = True
                            while no_keyword_confirm_input:
                                while no_input:
                                    with mic as source:
                                        tts.say(str("Do you want to see " + art_keyword + "?"))
                                        audio_text = recogniser.listen(source)
                                    try:
                                        keyword_confirm= recogniser.recognize_google(audio_text)
                                        if keyword_confirm == "yes":
                                            no_input = False
                                        elif keyword_confirm == "no":
                                            no_input = False

                                        else:
                                            tts.say("Sorry, I don't understand. Please try again.")
                                    except:
                                        tts.say("Sorry, can you repeat that?")
                                no_input = True
                                #confirms the input
                                while no_input:
                                    with mic as source:
                                        if keyword_confirm == "yes":
                                            tts.say(str("Are you sure that you want to see " + art_keyword + "?"))
                                        elif keyword_confirm == "no":
                                            tts.say(str("Are you sure that you don't want to see " + art_keyword + "?"))
                                        audio_text = recogniser.listen(source)
                                    try:
                                        confirm = recogniser.recognize_google(audio_text)
                                        if confirm == "yes":
                                            if keyword_confirm == "yes":
                                                no_input = False
                                                no_keyword_confirm_input = False
                                                no_room_input = False
                                                target = current_row[0].value
                                            else:
                                                no_input = False
                                                no_keyword_confirm_input = False
                                        elif confirm == "no":
                                            no_input = False
                                            no_keyword_confirm_input = False
                                        else:
                                            tts.say("Sorry, I don't understand. Please try again.")
                                    except:
                                        tts.say("Sorry, can you repeat that?")

                if no_room_input:
                    tts.say("Sorry I can't help you find that. Please try a different keyword.")
    return target

def find_instruct_type_input(mic,recogniser,tts):
    no_instruct_input = True
    while no_instruct_input:
        no_input = True
        while no_input:
            with mic as source:
                tts.say("Do you want to use a map?")
                audio_text = recogniser.listen(source)
            try:

                answer= recogniser.recognize_google(audio_text)
                if answer == "yes":
                    no_input = False
                elif answer == "no":
                    no_input = False
                else:
                    tts.say("Sorry, I don't understand. Please try again.")
            except:
                tts.say("Sorry, can you repeat that?")
        no_input = True
        while no_input:
            with mic as source:
                if answer == "yes":
                    tts.say(str("Are you sure that want to use a map?"))
                elif answer == "no":
                    tts.say(str("Are you sure that you don't want to use a map?"))
                audio_text = recogniser.listen(source)
            try:

                confirm = recogniser.recognize_google(audio_text)
                if confirm == "yes":
                    no_input = False
                    no_instruct_input = False
                elif confirm == "no":
                    no_input = False
                else:
                    tts.say("Sorry, I don't understand. Please try again.")
            except:
                tts.say("Sorry, can you repeat that?")
            if answer == "yes":
                instruct_type = "survey"
            elif answer == "no":
                instruct_type = "route"
    return instruct_type

def find_easy_access_input(mic,recogniser,tts):
    no_easy_access_input = True

    while no_easy_access_input:
        no_input = True
        while no_input:
            with mic as source:
                tts.say("Do you need to use wheelchair or pushchair access?")
                audio_text = recogniser.listen(source)
            try:

                easy_access= recogniser.recognize_google(audio_text)
                if easy_access == "yes":
                    no_input = False
                elif easy_access == "no":
                    no_input = False
                else:
                    tts.say("Sorry, I don't understand. Please try again.")
            except:
                tts.say("Sorry, can you repeat that?")
        no_input = True
        while no_input:
            with mic as source:
                if easy_access == "yes":
                    tts.say(str("Are you sure that you want to use wheelchair or pushchair access?"))
                elif easy_access == "no":
                    tts.say(str("Are you sure that you don't want to use wheelchair or pushchair access?"))
                audio_text = recogniser.listen(source)
            try:

                confirm = recogniser.recognize_google(audio_text)
                if confirm == "yes":
                    no_input = False
                    no_easy_access_input = False
                elif confirm == "no":
                    no_input = False
                else:
                    tts.say("Sorry, I don't understand. Please try again.")
            except:
                tts.say("Sorry, can you repeat that?")
    return easy_access

def main():
    ip_address, port_number, start = config_import()
    tts = ALProxy("ALTextToSpeech", str(ip_address), int(port_number))
    recogniser = speech_recognition.Recognizer()
    mic = speech_recognition.Microphone(device_index = 1)
    while True:
        target = find_room_input(mic,recogniser,tts)

        instruct_type = find_instruct_type_input(mic,recogniser,tts)

        easy_access = find_easy_access_input(mic,recogniser,tts)

        # special case due to theatre having easy access on floor below
        if target.lower() == "theatre" and easy_access == "yes":
            target = "access to theatre"

        road_network = map_import()
        #adjust road network for easy access
        road_network = map_adjust(road_network,easy_access)
        path = ftsp_alg.main(road_network,start.lower(),target.lower())
        if path == "failed":
            tts.say("Path not possible. Please try again")
        else:
            instructions_printer(start,path,road_network,instruct_type,tts,easy_access)

main()
